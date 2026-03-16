import json
import csv
import io
import openpyxl
from django.db.models import Q
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.contrib.auth import get_user_model
from django.urls import reverse
from accounts.utils import send_notification_email
from notifications.utils import create_notification
from .models import Exam, Question, ExamAttempt, StudentAnswer, Assignment, Class
from .forms import ExamForm, QuestionForm, AssignmentForm, BulkQuestionUploadForm

User = get_user_model()


@login_required
def exam_list(request):
    if request.user.user_type == 'student':
        student_classes = request.user.enrolled_classes.all()
        assignments = Assignment.objects.filter(
            Q(assigned_to_class__in=student_classes) | Q(assigned_to_students=request.user),
            is_active=True,
            start_date__lte=timezone.now(),
            end_date__gte=timezone.now()
        ).distinct()
        exams = Exam.objects.filter(assignments__in=assignments).distinct()
    else:
        exams = Exam.objects.filter(is_active=True)
    return render(request, 'exams/exam_list.html', {'exams': exams})


@login_required
def exam_detail(request, exam_id):
    """Sınav detayı ve başlama onayı"""
    exam = get_object_or_404(Exam, id=exam_id, is_active=True)
    attempt = ExamAttempt.objects.filter(student=request.user, exam=exam, is_finished=False).first()
    context = {
        'exam': exam,
        'attempt': attempt
    }
    return render(request, 'exams/exam_detail.html', context)


@login_required
def start_exam(request, exam_id):
    """Sınava başla, yeni bir attempt oluştur"""
    exam = get_object_or_404(Exam, id=exam_id, is_active=True)
    attempt, created = ExamAttempt.objects.get_or_create(
        student=request.user,
        exam=exam,
        is_finished=False,
        defaults={'start_time': timezone.now()}
    )
    return redirect('exams:take_exam', attempt_id=attempt.id)


@login_required
def take_exam(request, attempt_id):
    """Sınav çözme ekranı"""
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user, is_finished=False)
    exam = attempt.exam

    elapsed = (timezone.now() - attempt.start_time).total_seconds() / 60
    if elapsed > exam.duration:
        return redirect('exams:finish_exam', attempt_id=attempt.id)

    questions = exam.questions.all().order_by('order')
    answers_map = {str(ans.question_id): ans.selected_option for ans in attempt.answers.all()}

    context = {
        'attempt': attempt,
        'exam': exam,
        'questions': questions,
        'answers_map': answers_map,
        'remaining_time': int(exam.duration * 60 - (timezone.now() - attempt.start_time).total_seconds())
    }
    return render(request, 'exams/take_exam.html', context)


@require_POST
@login_required
def save_answer(request, attempt_id):
    """AJAX ile cevap kaydet"""
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user, is_finished=False)
    question_id = request.POST.get('question_id')
    selected = request.POST.get('selected')

    if not question_id:
        return HttpResponseBadRequest("Soru ID gerekli")

    question = get_object_or_404(Question, id=question_id, exam=attempt.exam)

    answer, created = StudentAnswer.objects.get_or_create(
        attempt=attempt,
        question=question,
        defaults={'selected_option': selected, 'is_correct': (selected == question.correct_answer)}
    )
    if not created:
        answer.selected_option = selected
        answer.is_correct = (selected == question.correct_answer)
        answer.save()

    return JsonResponse({'status': 'ok', 'saved': selected})


@login_required
def finish_exam(request, attempt_id):
    """Sınavı bitir ve puanı hesapla"""
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user, is_finished=False)
    exam = attempt.exam

    answers = attempt.answers.all()
    correct = answers.filter(is_correct=True).count()
    wrong = answers.filter(is_correct=False).exclude(selected_option__isnull=True).exclude(selected_option='').count()
    empty = exam.total_questions - (correct + wrong)

    attempt.correct_count = correct
    attempt.wrong_count = wrong
    attempt.empty_count = empty
    net = correct - (wrong / 4)
    attempt.score = net
    attempt.is_finished = True
    attempt.end_time = timezone.now()
    attempt.save()

    # Bildirimlerde kullanılacak net değer (2 ondalık)
    score_str = f"{attempt.score:.2f}"

    # Öğrenciye e-posta
    if request.user.email:
        context_student = {
            'student': request.user,
            'exam': exam,
            'attempt': attempt,
            'report_url': request.build_absolute_uri(
                reverse('reports:student_exam_detail', args=[attempt.id])
            ),
        }
        send_notification_email(
            subject='Sınav Tamamlandı',
            template_name='emails/exam_finished_student.html',
            context=context_student,
            recipient_list=[request.user.email],
        )

    # Öğretmene e-posta (sınavı oluşturan öğretmen)
    teacher = exam.created_by
    if teacher and teacher.email:
        context_teacher = {
            'teacher': teacher,
            'student': request.user,
            'exam': exam,
            'attempt': attempt,
            'report_url': request.build_absolute_uri(
                reverse('reports:teacher_student_detail', args=[request.user.id])
            ),
        }
        send_notification_email(
            subject='Öğrenci Sınavı Tamamladı',
            template_name='emails/exam_finished_teacher.html',
            context=context_teacher,
            recipient_list=[teacher.email],
        )

    # Veliye e-posta (eğer varsa)
    if getattr(request.user, 'parent', None) and request.user.parent.email:
        parent = request.user.parent
        context_parent = {
            'parent': parent,
            'student': request.user,
            'exam': exam,
            'attempt': attempt,
            'report_url': request.build_absolute_uri(
                reverse('reports:parent_child_exam_detail', args=[request.user.id, attempt.id])
            ),
        }
        send_notification_email(
            subject='Çocuğunuz Sınavı Tamamladı',
            template_name='emails/exam_finished_parent.html',
            context=context_parent,
            recipient_list=[parent.email],
        )

    # Öğrenciye bildirim
    create_notification(
        recipient=request.user,
        notification_type='exam_finished',
        title='Sınav Tamamlandı',
        message=f'{exam.title} sınavını tamamladınız. Net: {score_str}',
        link=reverse('reports:student_exam_detail', args=[attempt.id]),
        related_object=attempt,
    )

    # Öğretmene bildirim
    if teacher:
        create_notification(
            recipient=teacher,
            notification_type='exam_finished',
            title='Öğrenci Sınavı Tamamladı',
            message=(
                f'{request.user.get_full_name()} öğrenciniz '
                f'{exam.title} sınavını tamamladı. Net: {score_str}'
            ),
            link=reverse('reports:teacher_student_detail', args=[request.user.id]),
            sender=request.user,
            related_object=attempt,
        )

    # Veliye bildirim
    if getattr(request.user, 'parent', None):
        create_notification(
            recipient=request.user.parent,
            notification_type='exam_finished',
            title='Çocuğunuz Sınavı Tamamladı',
            message=(
                f'{request.user.get_full_name()} çocuğunuz '
                f'{exam.title} sınavını tamamladı. Net: {score_str}'
            ),
            link=reverse(
                'reports:parent_child_exam_detail',
                args=[request.user.id, attempt.id],
            ),
            related_object=attempt,
        )

    return redirect('exams:exam_result', attempt_id=attempt.id)


@login_required
def exam_result(request, attempt_id):
    """Sınav sonuç sayfası"""
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user, is_finished=True)
    exam = attempt.exam

    questions = exam.questions.all().order_by('order')
    results = []
    for q in questions:
        answer = attempt.answers.filter(question=q).first()
        results.append({
            'question': q,
            'selected': answer.selected_option if answer else None,
            'correct': q.correct_answer,
            'is_correct': answer.is_correct if answer else False
        })

    context = {
        'attempt': attempt,
        'exam': exam,
        'results': results,
    }
    return render(request, 'exams/exam_result.html', context)


# --- Öğretmen view'ları ---
def teacher_required(view_func):
    decorated = user_passes_test(
        lambda u: u.is_authenticated and getattr(u, 'user_type', None) == 'teacher',
        login_url='/accounts/giris/'
    )
    return decorated(view_func)


@teacher_required
def teacher_exam_list(request):
    exams = Exam.objects.filter(created_by=request.user)
    return render(request, 'exams/teacher/exam_list.html', {'exams': exams})


@teacher_required
def teacher_class_list(request):
    classes = Class.objects.filter(teacher=request.user)
    return render(request, 'exams/teacher/class_list.html', {'classes': classes})


@teacher_required
def teacher_class_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        student_ids = request.POST.getlist('students')
        if name:
            cls = Class.objects.create(name=name, teacher=request.user)
            if student_ids:
                cls.students.set(student_ids)
            messages.success(request, 'Sınıf oluşturuldu.')
            return redirect('exams:teacher_class_list')
    students = User.objects.filter(user_type='student')
    return render(request, 'exams/teacher/class_form.html', {'students': students})


@teacher_required
def teacher_class_manage(request, class_id):
    cls = get_object_or_404(Class, id=class_id, teacher=request.user)
    if request.method == 'POST':
        student_ids = request.POST.getlist('students')
        if student_ids:
            students = User.objects.filter(id__in=student_ids, user_type='student')
            cls.students.add(*students)
            messages.success(request, 'Öğrenciler sınıfa eklendi.')
        return redirect('exams:teacher_class_manage', class_id=cls.id)

    enrolled_students = cls.students.all().order_by('last_name', 'first_name')
    all_students = User.objects.filter(user_type='student').exclude(
        id__in=enrolled_students.values_list('id', flat=True)
    )
    return render(
        request,
        'exams/teacher/class_manage.html',
        {
            'class': cls,
            'enrolled_students': enrolled_students,
            'all_students': all_students,
        },
    )


@teacher_required
def teacher_class_remove_student(request, class_id, student_id):
    cls = get_object_or_404(Class, id=class_id, teacher=request.user)
    student = get_object_or_404(User, id=student_id, user_type='student')
    cls.students.remove(student)
    messages.success(request, f'{student.get_full_name()} sınıftan çıkarıldı.')
    return redirect('exams:teacher_class_manage', class_id=cls.id)


@teacher_required
def teacher_exam_create(request):
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            exam = form.save(commit=False)
            exam.created_by = request.user
            exam.save()
            messages.success(request, 'Sınav başarıyla oluşturuldu.')
            return redirect('exams:teacher_exam_list')
    else:
        form = ExamForm()
    return render(request, 'exams/teacher/exam_form.html', {'form': form, 'title': 'Sınav Oluştur'})


@teacher_required
def teacher_exam_edit(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, created_by=request.user)
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sınav güncellendi.')
            return redirect('exams:teacher_exam_list')
    else:
        form = ExamForm(instance=exam)
    return render(request, 'exams/teacher/exam_form.html', {'form': form, 'title': 'Sınav Düzenle'})


@teacher_required
def teacher_exam_delete(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, created_by=request.user)
    if request.method == 'POST':
        exam.delete()
        messages.success(request, 'Sınav silindi.')
        return redirect('exams:teacher_exam_list')
    return render(request, 'exams/teacher/exam_confirm_delete.html', {'exam': exam})


@teacher_required
def teacher_assign_exam(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, created_by=request.user)
    if request.method == 'POST':
        form = AssignmentForm(request.POST, teacher=request.user, exam=exam)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.exam = exam
            assignment.save()
            form.save_m2m()

            # E-posta ve bildirim için hedef öğrenciler
            if assignment.assigned_to_class:
                students = assignment.assigned_to_class.students.all()
            else:
                students = assignment.assigned_to_students.all()

            exam_url = request.build_absolute_uri(reverse('exams:detail', args=[exam.id]))

            for student in students:
                # E-posta
                if student.email:
                    context = {
                        'student': student,
                        'exam': exam,
                        'assignment': assignment,
                        'exam_url': exam_url,
                    }
                    send_notification_email(
                        subject='Yeni Sınav Atandı',
                        template_name='emails/exam_assigned.html',
                        context=context,
                        recipient_list=[student.email],
                    )

                # Site içi bildirim
                create_notification(
                    recipient=student,
                    notification_type='exam_assigned',
                    title='Yeni Sınav Atandı',
                    message=(
                        f'{exam.title} sınavı size atandı. '
                        f'Son tarih: {assignment.end_date.strftime("%d.%m.%Y %H:%M")}'
                    ),
                    link=reverse('exams:detail', args=[exam.id]),
                    sender=request.user,
                    related_object=assignment,
                )

            messages.success(request, 'Sınav başarıyla atandı.')
            return redirect('exams:teacher_exam_list')
    else:
        form = AssignmentForm(teacher=request.user, exam=exam)

    return render(request, 'exams/teacher/assign_exam.html', {'form': form, 'exam': exam})


@teacher_required
def teacher_bulk_question_upload(request):
    if request.method == 'POST':
        form = BulkQuestionUploadForm(request.POST, request.FILES, teacher=request.user)
        if form.is_valid():
            exam = form.cleaned_data['exam']
            file = form.cleaned_data['file']

            file_extension = file.name.split('.')[-1].lower()
            errors = []
            success_count = 0

            # CSV işleme – başlık satırına göre
            if file_extension == 'csv':
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                for row_num, row in enumerate(reader, start=2):  # 2. satırdan başla
                    try:
                        _create_question_from_row(row, exam)
                        success_count += 1
                    except Exception as e:
                        errors.append(f"Satır {row_num}: {str(e)}")

            # Excel işleme – ilk satırı başlık kabul et
            elif file_extension in ['xlsx', 'xls']:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    try:
                        _create_question_from_row(row_dict, exam)
                        success_count += 1
                    except Exception as e:
                        label = row[0] if row and row[0] else '?'
                        errors.append(f"Satır {label}: {str(e)}")

            else:
                errors.append("Desteklenmeyen dosya formatı. Lütfen CSV veya Excel yükleyin.")

            # Sonuç mesajları
            if success_count:
                messages.success(request, f"{success_count} soru başarıyla eklendi.")
            if errors:
                for err in errors[:10]:
                    messages.warning(request, err)
                if len(errors) > 10:
                    messages.warning(request, f"... ve {len(errors) - 10} hata daha.")

            return redirect('exams:teacher_bulk_question_upload')
    else:
        form = BulkQuestionUploadForm(teacher=request.user)

    return render(request, 'exams/teacher/bulk_upload.html', {'form': form})


@teacher_required
def teacher_bulk_upload(request, exam_id):
    """Belirli bir sınav için dosyadan toplu soru yükleme."""
    exam = get_object_or_404(Exam, id=exam_id, created_by=request.user)

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()

        errors = []
        success_count = 0

        # CSV
        if file_extension == 'csv':
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            for row_num, row in enumerate(reader, start=2):
                try:
                    _create_question_from_row(row, exam)
                    success_count += 1
                except Exception as e:
                    errors.append(f"Satır {row_num}: {str(e)}")

        # Excel
        elif file_extension in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                try:
                    _create_question_from_row(row_dict, exam)
                    success_count += 1
                except Exception as e:
                    label = row[0] if row and row[0] else '?'
                    errors.append(f"Satır {label}: {str(e)}")
        else:
            errors.append("Desteklenmeyen dosya formatı. Lütfen CSV veya Excel yükleyin.")

        if success_count:
            messages.success(request, f"{success_count} soru başarıyla eklendi.")
        if errors:
            for err in errors[:10]:
                messages.warning(request, err)
            if len(errors) > 10:
                messages.warning(request, f"... ve {len(errors) - 10} hata daha.")

        return redirect('exams:teacher_question_list', exam_id=exam.id)

    return render(request, 'exams/teacher/bulk_upload.html', {'exam': exam})


def _create_question_from_row(row, exam):
    """
    Satırdan soru oluşturur.

    Beklenen başlık örnekleri:
    - soru_metni / Soru Metni / soru
    - a, b, c, d, e veya secenek_a, secenek_b, ...
    - dogru_cevap / Doğru Cevap / cevap
    - sira / Sıra (opsiyonel)
    - aciklama / Açıklama (opsiyonel)
    """
    text = row.get('soru_metni') or row.get('Soru Metni') or row.get('soru')
    if not text:
        raise ValueError("Soru metni boş")

    option_a = row.get('a') or row.get('A') or row.get('secenek_a')
    option_b = row.get('b') or row.get('B') or row.get('secenek_b')
    option_c = row.get('c') or row.get('C') or row.get('secenek_c')
    option_d = row.get('d') or row.get('D') or row.get('secenek_d')
    option_e = row.get('e') or row.get('E') or row.get('secenek_e')

    if not all([option_a, option_b, option_c, option_d]):
        raise ValueError("A, B, C, D şıkları zorunludur")

    correct = row.get('dogru_cevap') or row.get('Doğru Cevap') or row.get('cevap')
    if not correct or str(correct).upper() not in ['A', 'B', 'C', 'D', 'E']:
        raise ValueError("Doğru cevap A,B,C,D,E olmalıdır")

    order_value = row.get('sira') or row.get('Sıra') or 0
    try:
        order = int(order_value)
    except (TypeError, ValueError):
        order = 0

    explanation = row.get('aciklama') or row.get('Açıklama') or ''

    Question.objects.create(
        exam=exam,
        text=text,
        option_a=option_a,
        option_b=option_b,
        option_c=option_c,
        option_d=option_d,
        option_e=option_e or None,
        correct_answer=str(correct).upper(),
        explanation=explanation or '',
        order=order,
    )


@teacher_required
def teacher_question_list(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, created_by=request.user)
    questions = exam.questions.all().order_by('order')
    return render(request, 'exams/teacher/question_list.html', {'exam': exam, 'questions': questions})


@teacher_required
def teacher_question_create(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, created_by=request.user)
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES)
        if form.is_valid():
            question = form.save(commit=False)
            question.exam = exam
            question.save()
            messages.success(request, 'Soru eklendi.')
            return redirect('exams:teacher_question_list', exam_id=exam.id)
    else:
        form = QuestionForm(initial={'exam': exam})
    return render(request, 'exams/teacher/question_form.html', {'form': form, 'exam': exam, 'title': 'Soru Ekle'})


@teacher_required
def teacher_question_edit(request, question_id):
    question = get_object_or_404(Question, id=question_id, exam__created_by=request.user)
    if request.method == 'POST':
        form = QuestionForm(request.POST, request.FILES, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'Soru güncellendi.')
            return redirect('exams:teacher_question_list', exam_id=question.exam.id)
    else:
        form = QuestionForm(instance=question)
    return render(request, 'exams/teacher/question_form.html', {'form': form, 'exam': question.exam, 'title': 'Soru Düzenle'})


@teacher_required
def teacher_question_delete(request, question_id):
    question = get_object_or_404(Question, id=question_id, exam__created_by=request.user)
    exam_id = question.exam.id
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Soru silindi.')
        return redirect('exams:teacher_question_list', exam_id=exam_id)
    return render(request, 'exams/teacher/question_confirm_delete.html', {'question': question})
